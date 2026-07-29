"""Helpers to turn extracted faculty/publication JSON records into clean
tabular data for on-screen tables and the two-sheet Excel export."""

import io
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

MASTER_COLUMN_ORDER = [
    "source_file", "name", "emails", "phones", "nationality", "country", "address",
    "current_designation", "current_department", "current_organization",
    # education_*_institution is a single merged field for each degree level
    # containing the combined institute/college and university name when both are present.
    "education_ug_degree", "education_ug_branch", "education_ug_institution", "education_ug_year",
    "education_pg_degree", "education_pg_branch", "education_pg_institution", "education_pg_year",
    "education_phd_degree", "education_phd_branch", "education_phd_institution", "education_phd_year",
    "experience_academic_years", "experience_industry_years", "experience_research_years",
    "experience_administrative_years", "experience_total_years", "experience_stated_total_years",
    "counts_journal", "counts_international_conference", "counts_national_conference",
    "counts_books", "counts_book_chapters", "counts_patents", "counts_projects",
    "counts_sponsored_projects", "counts_consultancy_projects", "counts_trainings", "counts_projects_general",
    "counts_awards", "counts_memberships", "counts_fdp_attended", "counts_sttp_attended",
    "counts_workshops_attended", "counts_online_courses",
    "profiles_orcid", "profiles_google_scholar", "profiles_researchgate", "profiles_linkedin", "profiles_scopus",
    "designation_history", "administrative_roles", "sponsored_projects", "consultancy_projects",
    "trainings_conducted", "projects_list", "awards_list", "memberships_list", "fdp_list", "sttp_list",
    "workshops_attended_list", "online_courses_list", "flags"
]

PUBLICATION_COLUMN_ORDER = [
    "source_file", "faculty_name", "sr_no", "title", "authors", "publication_type",
    "journal_or_conference", "publisher", "volume", "issue", "pages", "year", "month",
    "doi", "issn", "isbn", "indexed_in", "quartile", "impact_factor", "citation_count", "url",
]


def _join_list(value):
    if isinstance(value, list):
        return "; ".join(str(v) for v in value if v not in (None, ""))
    return value


def _join_records(value):
    if not isinstance(value, list):
        return value
    if not value:
        return ""
    if isinstance(value[0], dict):
        parts = []
        for d in value:
            title = d.get("title", "")
            date_range = d.get("date_range", "")
            if title and date_range:
                parts.append(f"{title} ({date_range})")
            elif title:
                parts.append(title)
        return "; ".join(parts)
    return "; ".join(str(v) for v in value if v not in (None, ""))


def _sanitize_missing_values(value):
    if value is None:
        return "NA"
    if isinstance(value, float) and pd.isna(value):
        return "NA"
    if value == "":
        return "NA"
    if value == [] or value == {}:
        return "NA"
    return value


def _sanitize_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    return df.apply(lambda col: col.map(_sanitize_missing_values))


def _flatten_pg_education(rec: dict) -> dict:
    """education.pg is now extracted as an ARRAY (to support multiple
    postgraduate-level degrees, e.g. an MCA earned before an M.Tech). For the
    flat Excel/table export, join multiple entries' degree/branch/institution
    /year into semicolon-separated strings within the SAME existing columns
    (education_pg_degree, etc.) — this keeps the export column structure
    backward-compatible while still surfacing every degree, rather than
    silently dropping all but one."""
    education = rec.get("education") or {}
    pg = education.get("pg")
    if isinstance(pg, list):
        degrees = "; ".join(p.get("degree", "") for p in pg if isinstance(p, dict) and p.get("degree"))
        branches = "; ".join(p.get("branch", "") for p in pg if isinstance(p, dict) and p.get("branch"))
        institutions = "; ".join(p.get("institution", "") for p in pg if isinstance(p, dict) and p.get("institution"))
        years = "; ".join(str(p.get("year", "")) for p in pg if isinstance(p, dict) and p.get("year"))
        education = dict(education)
        education["pg"] = {"degree": degrees, "branch": branches, "institution": institutions, "year": years}
        rec = dict(rec)
        rec["education"] = education
    return rec


def master_records_to_dataframe(records: list) -> pd.DataFrame:
    """Flatten nested master_faculty_database records into a flat DataFrame."""
    if not records:
        return pd.DataFrame(columns=MASTER_COLUMN_ORDER)

    # Join list fields before normalizing so they render as plain strings, not
    # python list reprs, in both the on-screen table and the Excel sheet.
    prepped = []
    for rec in records:
        rec = _flatten_pg_education(rec)
        rec = dict(rec)
        rec["emails"] = _join_records(rec.get("emails"))
        rec["phones"] = _join_records(rec.get("phones"))
        rec["designation_history"] = _join_records(rec.get("designation_history"))
        rec["administrative_roles"] = _join_records(rec.get("administrative_roles"))
        rec["sponsored_projects"] = _join_records(rec.get("sponsored_projects"))
        rec["consultancy_projects"] = _join_records(rec.get("consultancy_projects"))
        rec["trainings_conducted"] = _join_records(rec.get("trainings_conducted"))
        rec["projects_list"] = _join_records(rec.get("projects_list"))
        rec["awards_list"] = _join_records(rec.get("awards_list"))
        rec["memberships_list"] = _join_records(rec.get("memberships_list"))
        rec["fdp_list"] = _join_records(rec.get("fdp_list"))
        rec["sttp_list"] = _join_records(rec.get("sttp_list"))
        rec["workshops_attended_list"] = _join_records(rec.get("workshops_attended_list"))
        rec["online_courses_list"] = _join_records(rec.get("online_courses_list"))
        rec["flags"] = _join_records(rec.get("flags"))
        prepped.append(rec)

    df = pd.json_normalize(prepped, sep="_")

    # Ensure every expected column exists, in a stable, readable order.
    for col in MASTER_COLUMN_ORDER:
        if col not in df.columns:
            df[col] = None
    extra_cols = [c for c in df.columns if c not in MASTER_COLUMN_ORDER]
    df = df[MASTER_COLUMN_ORDER + extra_cols]
    return _sanitize_dataframe(df)


def publication_records_to_dataframe(records: list) -> pd.DataFrame:
    """Flatten publication_details records into a flat DataFrame."""
    if not records:
        return pd.DataFrame(columns=PUBLICATION_COLUMN_ORDER)

    prepped = []
    for rec in records:
        rec = dict(rec)
        rec["authors"] = _join_list(rec.get("authors"))
        rec["indexed_in"] = _join_list(rec.get("indexed_in"))
        prepped.append(rec)

    df = pd.json_normalize(prepped, sep="_")

    for col in PUBLICATION_COLUMN_ORDER:
        if col not in df.columns:
            df[col] = None
    extra_cols = [c for c in df.columns if c not in PUBLICATION_COLUMN_ORDER]
    df = df[PUBLICATION_COLUMN_ORDER + extra_cols]
    return _sanitize_dataframe(df)


PUBLICATION_LAYOUT_COLUMNS = [
    "sr_no", "title", "authors", "publication_type",
    "journal_or_conference", "publisher", "volume", "issue", "pages",
    "year", "month", "doi", "issn", "isbn", "indexed_in",
    "quartile", "impact_factor", "citation_count", "url",
]

PUBLICATION_LAYOUT_HEADERS = [
    "Sr No", "Title", "Authors", "Publication Type", "Journal/Conference",
    "Publisher", "Volume", "Issue", "Pages", "Year", "Month", "DOI",
    "ISSN", "ISBN", "Indexed In", "Quartile", "Impact Factor",
    "Citation Count", "URL",
]


def _autofit_columns(worksheet: openpyxl.worksheet.worksheet.Worksheet):
    width_map = {}
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, max_col=len(PUBLICATION_LAYOUT_HEADERS)):
        for cell in row:
            if cell.value is None:
                continue
            value = str(cell.value)
            column_letter = cell.column_letter
            width_map[column_letter] = max(width_map.get(column_letter, 0), len(value))
    for column_letter, width in width_map.items():
        worksheet.column_dimensions[column_letter].width = min(max(12, width + 2), 60)


def _write_publication_sheet_grouped(worksheet: openpyxl.worksheet.worksheet.Worksheet, master_df: pd.DataFrame, publication_df: pd.DataFrame):
    label_font = Font(bold=True)
    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="FFDDDDDD")
    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    current_row = 1
    grouped = publication_df.groupby("source_file", sort=False)

    if master_df.empty:
        worksheet.cell(row=current_row, column=1, value="No faculty records available").font = label_font
        return

    for idx, master_row in master_df.iterrows():
        source_file = master_row.get("source_file", "")
        faculty_name = master_row.get("name", "")
        designation = master_row.get("current_designation", "")
        organization = master_row.get("current_organization", "")

        worksheet.cell(row=current_row, column=1, value="Name:").font = label_font
        worksheet.cell(row=current_row, column=2, value=faculty_name)
        current_row += 1

        worksheet.cell(row=current_row, column=1, value="Designation:").font = label_font
        worksheet.cell(row=current_row, column=2, value=designation)
        current_row += 1

        worksheet.cell(row=current_row, column=1, value="Organization:").font = label_font
        worksheet.cell(row=current_row, column=2, value=organization)
        current_row += 2

        publications = grouped.get_group(source_file) if source_file in grouped.groups else None
        if publications is None or publications.empty:
            worksheet.cell(row=current_row, column=1, value="No publications recorded")
            current_row += 1
        else:
            for col_idx, header in enumerate(PUBLICATION_LAYOUT_HEADERS, start=1):
                cell = worksheet.cell(row=current_row, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = wrap_alignment
            current_row += 1

            for _, publication in publications.iterrows():
                for col_idx, key in enumerate(PUBLICATION_LAYOUT_COLUMNS, start=1):
                    worksheet.cell(row=current_row, column=col_idx, value=publication.get(key, ""))
                current_row += 1

        current_row += 2

    _autofit_columns(worksheet)


def build_publication_sheet_grouped(master_df: pd.DataFrame, publication_df: pd.DataFrame, sheet_name: str = "Publication Details") -> bytes:
    buffer = io.BytesIO()
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    _write_publication_sheet_grouped(worksheet, master_df, publication_df)
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


FACULTY_SHEET_HEADERS = [
    "Sr No", "Publication Type", "Title of Paper", "Journal Name",
    "Published Under / Journal Details", "Publisher", "Year of Publication",
    "DOI", "ISSN / ISBN", "Scopus Index",
]

INVALID_SHEET_NAME_CHARS = ["\\", "/", "*", "?", ":", "[", "]"]


def _safe_sheet_name(raw_name: str, used_names: dict) -> str:
    """Turn a faculty name / source_file into a valid, unique worksheet name.

    Excel sheet names must be <=31 chars and cannot contain \\ / * ? : [ ].
    Duplicates are disambiguated as 'Name', 'Name (2)', 'Name (3)', ...,
    matching the numbering style used in the target export.
    """
    is_missing = raw_name in (None, "") or (isinstance(raw_name, float) and pd.isna(raw_name))
    name = "Unknown" if is_missing else str(raw_name).strip()
    name = name or "Unknown"
    for ch in INVALID_SHEET_NAME_CHARS:
        name = name.replace(ch, "-")
    name = name[:31] or "Unknown"

    count = used_names.get(name, 0) + 1
    used_names[name] = count
    if count == 1:
        return name
    suffix = f" ({count})"
    return name[: 31 - len(suffix)] + suffix


def _write_faculty_publication_sheet(worksheet: openpyxl.worksheet.worksheet.Worksheet, master_row, publications: pd.DataFrame):
    label_font = Font(bold=True)
    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="FFDDDDDD")
    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    worksheet.cell(row=1, column=1, value="Name of Faculty :").font = label_font
    worksheet.cell(row=1, column=2, value=master_row.get("name", ""))

    worksheet.cell(row=2, column=1, value="Designation :").font = label_font
    worksheet.cell(row=2, column=2, value=master_row.get("current_designation", ""))

    worksheet.cell(row=3, column=1, value="Branch :").font = label_font
    worksheet.cell(row=3, column=2, value=master_row.get("current_department", ""))

    current_row = 5
    for col_idx, header in enumerate(FACULTY_SHEET_HEADERS, start=1):
        cell = worksheet.cell(row=current_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap_alignment
    current_row += 1

    if publications is None or publications.empty:
        worksheet.cell(row=current_row, column=1, value="No publications recorded")
    else:
        for _, pub in publications.iterrows():
            issn_isbn = "; ".join(v for v in [pub.get("issn", ""), pub.get("isbn", "")] if v not in (None, "", "NA"))
            row_values = [
                pub.get("sr_no", ""),
                pub.get("publication_type", ""),
                pub.get("title", ""),
                pub.get("journal_or_conference", ""),
                pub.get("authors", ""),
                pub.get("publisher", ""),
                pub.get("year", ""),
                pub.get("doi", ""),
                issn_isbn,
                pub.get("indexed_in", ""),
            ]
            for col_idx, value in enumerate(row_values, start=1):
                worksheet.cell(row=current_row, column=col_idx, value=value)
            current_row += 1

    _autofit_columns(worksheet)


def build_publication_workbook_per_faculty(master_df: pd.DataFrame, publication_df: pd.DataFrame) -> bytes:
    """Build a workbook with ONE WORKSHEET PER FACULTY/FILE, matching the
    'Name of Faculty / Designation / Branch' header layout followed by a
    publication table, instead of stacking every faculty into one sheet."""
    buffer = io.BytesIO()
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    if master_df.empty:
        sheet = workbook.create_sheet(title="No Data")
        sheet.cell(row=1, column=1, value="No faculty records available")
        workbook.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    grouped = publication_df.groupby("source_file", sort=False) if not publication_df.empty else None
    used_names = {}

    for _, master_row in master_df.iterrows():
        source_file = master_row.get("source_file", "")
        name_val = master_row.get("name", "")
        name_is_missing = name_val in (None, "") or (isinstance(name_val, float) and pd.isna(name_val))
        faculty_name = source_file if name_is_missing else name_val
        faculty_name = faculty_name if faculty_name not in (None, "") else "Unknown"
        sheet_name = _safe_sheet_name(faculty_name, used_names)
        sheet = workbook.create_sheet(title=sheet_name)

        publications = None
        if grouped is not None and source_file in grouped.groups:
            publications = grouped.get_group(source_file)

        _write_faculty_publication_sheet(sheet, master_row, publications)

    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def build_excel_workbook(master_df: pd.DataFrame, publication_df: pd.DataFrame) -> bytes:
    """Build a two-sheet .xlsx workbook in memory and return its bytes."""
    buffer = io.BytesIO()
    workbook = openpyxl.Workbook()
    master_sheet = workbook.active
    master_sheet.title = "Master Faculty Database"

    for col_idx, col_name in enumerate(master_df.columns, start=1):
        master_sheet.cell(row=1, column=col_idx, value=col_name).font = Font(bold=True)

    for row_idx, row in enumerate(master_df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            master_sheet.cell(row=row_idx, column=col_idx, value=value)

    _autofit_columns(master_sheet)

    publication_sheet = workbook.create_sheet(title="Publication Details")
    _write_publication_sheet_grouped(publication_sheet, master_df, publication_df)

    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def build_single_sheet_workbook(df: pd.DataFrame, sheet_name: str) -> bytes:
    """Build a single-sheet .xlsx workbook in memory and return its bytes."""
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        worksheet = writer.sheets[sheet_name]
        for i, col in enumerate(df.columns, start=1):
            max_len = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str).tolist()])
            worksheet.column_dimensions[worksheet.cell(row=1, column=i).column_letter].width = min(max(12, max_len + 2), 60)
    buffer.seek(0)
    return buffer.getvalue()
